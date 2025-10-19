"""
Memory-efficient Excel processing utilities.
Replaces Pandas with streaming alternatives for large datasets.
"""

import csv
import logging
import os
from typing import List, Dict, Any, Iterator
from datetime import datetime
import gc

logger = logging.getLogger(__name__)

class StreamingExcelExporter:
    """Memory-efficient Excel exporter using openpyxl streaming."""
    
    def __init__(self, chunk_size: int = 1000):
        self.chunk_size = chunk_size
        
    def export_cases_to_excel_streaming(
        self,
        cases: Iterator[Dict],
        filepath: str,
        sheet_name: str = "Cases",
    ) -> str:
        """Export cases to Excel using streaming to avoid memory issues."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import NamedStyle
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Define headers
            headers = [
                "Case No", "Date Reported", "Category", "Amount", 
                "List", "Status", "Responsibility", "Description"
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Style headers
            header_style = NamedStyle(name="header")
            header_style.font.bold = True
            for col in range(1, len(headers) + 1):
                ws.cell(row=1, column=col).style = header_style
            
            row_num = 2
            total_amount = 0.0
            case_count = 0
            
            # Process cases in chunks
            chunk = []
            for case in cases:
                chunk.append(case)
                
                if len(chunk) >= self.chunk_size:
                    self._write_chunk_to_worksheet(ws, chunk, row_num, headers)
                    row_num += len(chunk)
                    case_count += len(chunk)
                    total_amount += sum(float(c.get('amount', 0)) for c in chunk)
                    chunk = []
                    gc.collect()  # Free memory after each chunk
            
            # Write remaining cases
            if chunk:
                self._write_chunk_to_worksheet(ws, chunk, row_num, headers)
                case_count += len(chunk)
                total_amount += sum(float(c.get('amount', 0)) for c in chunk)
            
            # Add summary information
            self._add_summary_to_worksheet(ws, case_count, total_amount, len(headers))
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws, headers)
            
            # Save file
            wb.save(filepath)
            logger.info("Exported %s cases to %s", case_count, filepath)
            
            return filepath
            
        except ImportError:
            # Fallback to CSV if openpyxl not available
            return self._export_to_csv_fallback(cases, filepath)
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def _write_chunk_to_worksheet(self, ws, chunk: List[Dict], start_row: int, headers: List[str]):
        """Write a chunk of cases to worksheet."""
        for i, case in enumerate(chunk):
            row = start_row + i
            
            # Map case data to columns
            values = [
                case.get('transaction_no', ''),
                case.get('date_reported', ''),
                case.get('category', ''),
                case.get('amount', 0),
                case.get('list', ''),
                case.get('status', ''),
                case.get('responsibility', ''),
                case.get('description', '')
            ]
            
            for col, value in enumerate(values, 1):
                ws.cell(row=row, column=col, value=value)
    
    def _add_summary_to_worksheet(self, ws, case_count: int, total_amount: float, num_columns: int):
        """Add summary information to worksheet."""
        from openpyxl.utils import get_column_letter
        
        # Insert rows at the top
        ws.insert_rows(1, 4)
        
        ws['A1'] = "Cases Export Summary"
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"Total Cases: {case_count}"
        ws['A4'] = f"Total Amount: R {total_amount:,.2f}"
        
        # Merge cells for title
        last_col = get_column_letter(num_columns)
        ws.merge_cells(f'A1:{last_col}1')
    
    def _auto_adjust_columns(self, ws, headers: List[str]):
        """Auto-adjust column widths."""
        from openpyxl.utils import get_column_letter
        
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except Exception:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _export_to_csv_fallback(self, cases: Iterator[Dict], filepath: str) -> str:
        """Fallback CSV export if Excel libraries not available."""
        csv_path = filepath.replace('.xlsx', '.csv')
        
        headers = [
            "Case No", "Date Reported", "Category", "Amount", 
            "List", "Status", "Responsibility", "Description"
        ]
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            
            for case in cases:
                writer.writerow({
                    'Case No': case.get('transaction_no', ''),
                    'Date Reported': case.get('date_reported', ''),
                    'Category': case.get('category', ''),
                    'Amount': case.get('amount', 0),
                    'List': case.get('list', ''),
                    'Status': case.get('status', ''),
                    'Responsibility': case.get('responsibility', ''),
                    'Description': case.get('description', '')
                })
        
        logger.info("Exported to CSV fallback: %s", csv_path)
        return csv_path

class OptimizedReportGenerator:
    """Memory-efficient report generator using SQL aggregations."""
    
    def __init__(self):
        pass
    
    def generate_case_summary_report(self, fy_id: int = None) -> Dict[str, Any]:
        """Generate case summary using efficient SQL queries."""
        from scripts.Utilities.db_utils import get_db_connection
        
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                # Base query conditions
                where_clause = "WHERE 1=1"
                params = []
                
                if fy_id:
                    where_clause += " AND fy_id = ?"
                    params.append(fy_id)
                
                # Get total cases and amount
                cursor.execute(f"""
                    SELECT 
                        COUNT(*) as total_cases,
                        SUM(amount) as total_amount,
                        AVG(amount) as avg_amount
                    FROM cases 
                    {where_clause}
                    AND list != 'Deleted Cases'
                """, params)
                
                totals = cursor.fetchone()
                
                # Get cases by category
                cursor.execute(f"""
                    SELECT 
                        category,
                        COUNT(*) as count,
                        SUM(amount) as total_amount
                    FROM cases 
                    {where_clause}
                    AND list != 'Deleted Cases'
                    GROUP BY category
                    ORDER BY total_amount DESC
                """, params)
                
                categories = cursor.fetchall()
                
                # Get cases by status
                cursor.execute(f"""
                    SELECT 
                        assessment_status,
                        COUNT(*) as count,
                        SUM(amount) as total_amount
                    FROM cases 
                    {where_clause}
                    AND list != 'Deleted Cases'
                    GROUP BY assessment_status
                    ORDER BY total_amount DESC
                """, params)
                
                statuses = cursor.fetchall()
                
                # Get cases by list
                cursor.execute(f"""
                    SELECT 
                        list,
                        COUNT(*) as count,
                        SUM(amount) as total_amount
                    FROM cases 
                    {where_clause}
                    AND list != 'Deleted Cases'
                    GROUP BY list
                    ORDER BY total_amount DESC
                """, params)
                
                lists = cursor.fetchall()
                
                return {
                    'totals': {
                        'total_cases': totals[0] or 0,
                        'total_amount': totals[1] or 0.0,
                        'avg_amount': totals[2] or 0.0
                    },
                    'categories': [
                        {'category': cat[0], 'count': cat[1], 'amount': cat[2]}
                        for cat in categories
                    ],
                    'statuses': [
                        {'status': stat[0], 'count': stat[1], 'amount': stat[2]}
                        for stat in statuses
                    ],
                    'lists': [
                        {'list': lst[0], 'count': lst[1], 'amount': lst[2]}
                        for lst in lists
                    ]
                }
                
        except Exception as e:
            logger.error(f"Error generating report: {e}")
            raise
    
    def generate_responsibility_report(self, fy_id: int = None) -> Dict[str, Any]:
        """Generate responsibility-based report using SQL aggregations."""
        from scripts.Utilities.db_utils import get_db_connection
        
        try:
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                where_clause = "WHERE c.list != 'Deleted Cases'"
                params = []
                
                if fy_id:
                    where_clause += " AND c.fy_id = ?"
                    params.append(fy_id)
                
                cursor.execute(f"""
                    SELECT 
                        r.name as responsibility_name,
                        COUNT(c.id) as case_count,
                        SUM(c.amount) as total_amount,
                        AVG(c.amount) as avg_amount
                    FROM cases c
                    JOIN responsibilities r ON c.responsibility_id = r.id
                    {where_clause}
                    GROUP BY r.id, r.name
                    ORDER BY total_amount DESC
                    LIMIT 50
                """, params)
                
                responsibilities = cursor.fetchall()
                
                return {
                    'responsibilities': [
                        {
                            'name': resp[0],
                            'case_count': resp[1],
                            'total_amount': resp[2],
                            'avg_amount': resp[3]
                        }
                        for resp in responsibilities
                    ]
                }
                
        except Exception as e:
            logger.error(f"Error generating responsibility report: {e}")
            raise

def create_optimized_excel_export(dialog, cases_data: List[Dict], filename: str) -> str:
    """Create optimized Excel export with memory management."""
    try:
        # Create year folder
        from scripts.Utilities.financial_utils import create_year_folder, get_financial_year
        year_folder = create_year_folder(get_financial_year())
        export_dir = os.path.join(year_folder, "Exports")
        os.makedirs(export_dir, exist_ok=True)
        
        filepath = os.path.join(export_dir, filename)
        
        # Use streaming exporter
        exporter = StreamingExcelExporter(chunk_size=1000)
        
        # Convert list to iterator for memory efficiency
        def cases_iterator():
            for case in cases_data:
                yield case
        
        return exporter.export_cases_to_excel_streaming(
            cases_iterator(), 
            filepath, 
            "Cases Export"
        )
        
    except Exception as e:
        logger.error(f"Error in optimized Excel export: {e}")
        raise

def validate_memory_for_export(cases_count: int) -> bool:
    """Validate if there's enough memory for export operation."""
    try:
        import psutil
        
        available_memory_gb = psutil.virtual_memory().available / (1024**3)
        
        # Estimate memory needed (rough calculation)
        estimated_memory_mb = cases_count * 0.001  # ~1KB per case
        
        threshold_mb = available_memory_gb * 1024 * 0.5
        if estimated_memory_mb > threshold_mb:  # Use max 50% of available memory
            logger.warning(
                "Large export detected: %s cases, %.1fMB estimated",
                cases_count,
                estimated_memory_mb,
            )
            return False
        
        return True
        
    except ImportError:
        # psutil not available, assume OK for small exports
        return cases_count < 10000
