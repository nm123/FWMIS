import json
import os
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from scripts.Utilities.annexure_utils import get_annexure_details


def export_annexure_to_excel(annexure_ids: List[Optional[int]], file_path: str):
    """Export annexures to Excel format."""
    workbook = Workbook()

    # Remove default sheet
    workbook.remove(workbook.active)

    # Process each annexure
    for annexure_id in annexure_ids:
        if annexure_id:
            annexure = get_annexure_details(annexure_id)
            if annexure:
                create_annexure_sheet(workbook, annexure)

    # Save workbook
    workbook.save(file_path)


def create_annexure_sheet(workbook: Workbook, annexure: dict):
    """Create a worksheet for a single annexure."""
    sheet_name = f"{annexure['annexure_no']} ({annexure['role']})"
    # Excel sheet names can't exceed 31 characters
    if len(sheet_name) > 31:
        sheet_name = sheet_name[:31]

    worksheet = workbook.create_sheet(title=sheet_name)

    # Header styling
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Data styling
    data_font = Font(size=11)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Border styling
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title
    worksheet.merge_cells("A1:F1")
    title_cell = worksheet["A1"]
    title_cell.value = f"WRITE-OFF ANNEXURE: {annexure['annexure_no']}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(
        start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
    )

    # Subtitle
    worksheet.merge_cells("A2:F2")
    subtitle_cell = worksheet["A2"]
    subtitle_cell.value = f"Approval Authority: {annexure['role']}"
    subtitle_cell.font = Font(bold=True, size=12)
    subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Summary information
    worksheet["A4"] = "Total Cases:"
    worksheet["B4"] = annexure["case_count"]
    worksheet["D4"] = "Total Amount:"
    worksheet["E4"] = f"R {annexure['total_amount']:,.2f}"

    # Format summary row
    for col in ["A", "B", "D", "E"]:
        cell = worksheet[f"{col}4"]
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )

    # Column headers
    headers = [
        "Case No",
        "Responsibility",
        "Amount",
        "Description",
        "LC Recommendation",
        "LC Minutes",
    ]
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=6, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

        # Set column width
        if col == 1:  # Case No
            worksheet.column_dimensions[get_column_letter(col)].width = 15
        elif col == 2:  # Responsibility
            worksheet.column_dimensions[get_column_letter(col)].width = 20
        elif col == 3:  # Amount
            worksheet.column_dimensions[get_column_letter(col)].width = 15
        elif col == 4:  # Description
            worksheet.column_dimensions[get_column_letter(col)].width = 40
        elif col == 5:  # LC Recommendation
            worksheet.column_dimensions[get_column_letter(col)].width = 20
        elif col == 6:  # LC Minutes
            worksheet.column_dimensions[get_column_letter(col)].width = 20

    # Data rows
    for row_idx, case in enumerate(annexure["cases"], 7):
        # Case No
        case_cell = worksheet.cell(row=row_idx, column=1, value=case["transaction_no"])
        case_cell.font = data_font
        case_cell.alignment = data_alignment
        case_cell.border = thin_border

        # Responsibility
        resp_cell = worksheet.cell(
            row=row_idx, column=2, value=case["responsibility_name"]
        )
        resp_cell.font = data_font
        resp_cell.alignment = data_alignment
        resp_cell.border = thin_border

        # Amount
        amount_cell = worksheet.cell(
            row=row_idx, column=3, value=f"R {case['amount']:,.2f}"
        )
        amount_cell.font = data_font
        amount_cell.alignment = Alignment(horizontal="right", vertical="center")
        amount_cell.border = thin_border

        # Description
        desc_cell = worksheet.cell(row=row_idx, column=4, value=case["description"])
        desc_cell.font = data_font
        desc_cell.alignment = data_alignment
        desc_cell.border = thin_border

        # LC Recommendation
        lc_rec_cell = worksheet.cell(
            row=row_idx, column=5, value="Write-Off Recommended"
        )
        lc_rec_cell.font = data_font
        lc_rec_cell.alignment = data_alignment
        lc_rec_cell.border = thin_border

        # LC Minutes
        lc_minutes_path = get_lc_minutes_path(case["evidence_paths"])
        if lc_minutes_path and os.path.exists(lc_minutes_path):
            # Create hyperlink to LC minutes file
            lc_cell = worksheet.cell(row=row_idx, column=6, value="View LC Minutes")
            lc_cell.hyperlink = Hyperlink(
                target=lc_minutes_path, tooltip=lc_minutes_path
            )
            lc_cell.font = Font(size=11, color="0000FF", underline="single")
        else:
            lc_cell = worksheet.cell(row=row_idx, column=6, value="Not Available")
            lc_cell.font = Font(size=11, color="FF0000")

        lc_cell.alignment = data_alignment
        lc_cell.border = thin_border

    # Add totals row
    total_row = len(annexure["cases"]) + 8
    worksheet.cell(row=total_row, column=2, value="TOTAL:")
    worksheet.cell(row=total_row, column=3, value=f"R {annexure['total_amount']:,.2f}")

    # Format totals row
    for col in ["B", "C"]:
        cell = worksheet[f"{col}{total_row}"]
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(
            start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"
        )
        cell.border = thin_border

    # Set row heights
    worksheet.row_dimensions[1].height = 30  # Title
    worksheet.row_dimensions[2].height = 25  # Subtitle
    worksheet.row_dimensions[6].height = 25  # Headers

    # Freeze panes
    worksheet.freeze_panes = "A7"


def get_lc_minutes_path(evidence_paths: str) -> Optional[str]:
    """Extract LC minutes file path from evidence_paths JSON."""
    if not evidence_paths:
        return None

    try:
        evidence_data = json.loads(evidence_paths)
        lc_minutes = evidence_data.get("lc_minutes") or evidence_data.get(
            "loss_control_minutes"
        )
        if lc_minutes and isinstance(lc_minutes, str):
            return lc_minutes
        return None
    except (json.JSONDecodeError, TypeError):
        return None
