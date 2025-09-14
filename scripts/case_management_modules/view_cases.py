import os
import sqlite3
from datetime import datetime
from functools import partial
from PyQt5.QtWidgets import (
    QDialog,
    QVBoxLayout,
    QHBoxLayout,
    QTreeWidget,
    QTreeWidgetItem,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QMessageBox,
    QSplitter,
    QWidget,
    QLabel,
    QLineEdit,
    QScrollArea,
    QFormLayout,
    QGroupBox,
    QTextEdit,
    QComboBox,
    QPushButton,
)
from PyQt5.QtCore import Qt, QEvent
from PyQt5.QtGui import QWheelEvent
from scripts.Utilities.config import DB_PATH
from scripts.Utilities.utils import format_currency_amount
from scripts.Utilities.responsibility_utils import load_responsibilities
from scripts.Utilities.tree_utils import get_subtree_resp_ids
from scripts.Utilities.ui_theme import apply_theme, create_professional_button
from scripts.Utilities.financial_utils import get_all_financial_years, get_current_open_financial_year, get_financial_year
from collections import defaultdict


class NoWheelComboBox(QComboBox):
    """Custom QComboBox that ignores mouse wheel events unless focused"""

    def wheelEvent(self, event: QWheelEvent):
        """Override wheel event to only accept when widget has focus"""
        if self.hasFocus():
            super().wheelEvent(event)
        else:
            # Ignore wheel event when not focused
            event.ignore()


class ViewCasesDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("View Cases")
        self.setFixedSize(1700, 600)  # Increased by another ~10% (160px) for optimal header visibility
        self.responsibilities = load_responsibilities()

        # Apply professional theme
        apply_theme(self)

        self.setup_ui()

    def populate_fy_filter(self):
        """Populate the financial year filter combo box"""
        self.fy_filter_combo.clear()

        # Get all financial years
        financial_years = get_all_financial_years()

        # Add financial years to combo box
        for fy_id, fy_string, is_open in financial_years:
            self.fy_filter_combo.addItem(fy_string, fy_id)

        # Set current open financial year as default
        current_open = get_current_open_financial_year()
        if current_open:
            fy_id, fy_string = current_open
            index = self.fy_filter_combo.findData(fy_id)
            if index >= 0:
                self.fy_filter_combo.setCurrentIndex(index)

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Compact search bars layout
        search_layout = QHBoxLayout()
        search_layout.setContentsMargins(5, 5, 5, 5)
        search_layout.setSpacing(10)

        # Financial Year filter
        fy_label = QLabel("FY:")
        fy_label.setFixedWidth(20)
        self.fy_filter_combo = NoWheelComboBox()
        self.fy_filter_combo.setFixedWidth(120)
        self.populate_fy_filter()
        self.fy_filter_combo.currentTextChanged.connect(lambda: self.refresh_cases())

        search_layout.addWidget(fy_label)
        search_layout.addWidget(self.fy_filter_combo)

        # Separator
        search_layout.addSpacing(20)

        # Responsibility search
        resp_label = QLabel("Responsibility:")
        resp_label.setFixedWidth(80)
        self.resp_search_edit = QLineEdit()
        self.resp_search_edit.setPlaceholderText("Type to search...")
        self.resp_search_edit.setFixedWidth(200)
        self.resp_search_edit.textChanged.connect(self.filter_responsibilities)

        search_layout.addWidget(resp_label)
        search_layout.addWidget(self.resp_search_edit)

        # Separator
        search_layout.addSpacing(20)

        # List filter - compact layout
        list_label = QLabel("List:")
        list_label.setFixedWidth(30)
        self.list_filter_combo = NoWheelComboBox()
        self.list_filter_combo.addItems([
            "All Cases", "Checklist", "Lead Schedule", "To-Do List",
            "Recovered", "Write-Off Recommended", "Written Off", "Deleted Cases"
        ])
        self.list_filter_combo.setCurrentText("All Cases")
        self.list_filter_combo.setFixedWidth(140)  # Increased width for longer list names
        self.list_filter_combo.currentTextChanged.connect(lambda: (self.refresh_cases(), self.update_write_off_buttons_visibility()))

        search_layout.addWidget(list_label)
        search_layout.addWidget(self.list_filter_combo)

        search_layout.addStretch()
        layout.addLayout(search_layout)

        # Write-Off Recommended specific buttons (shown only when in that list)
        self.write_off_buttons_layout = QHBoxLayout()
        self.write_off_buttons_layout.setContentsMargins(5, 0, 5, 10)
        self.write_off_buttons_layout.setSpacing(10)

        self.create_submission_btn = create_professional_button("Create Write-Off Submission", "primary")
        self.create_submission_btn.clicked.connect(self.create_write_off_submission)
        self.create_submission_btn.setVisible(False)  # Hidden by default
        self.write_off_buttons_layout.addWidget(self.create_submission_btn)

        self.approve_submission_btn = create_professional_button("Approve Write-Off Submission", "success")
        self.approve_submission_btn.clicked.connect(self.approve_write_off_submission)
        self.approve_submission_btn.setVisible(False)  # Hidden by default
        self.write_off_buttons_layout.addWidget(self.approve_submission_btn)

        # Excel export button (available for all lists)
        self.excel_export_btn = create_professional_button("Export to Excel", "info")
        self.excel_export_btn.clicked.connect(self.export_to_excel)
        self.write_off_buttons_layout.addWidget(self.excel_export_btn)

        self.write_off_buttons_layout.addStretch()
        layout.addLayout(self.write_off_buttons_layout)

        # Main content layout
        content_layout = QHBoxLayout()
        splitter = QSplitter(Qt.Horizontal)

        self.resp_tree = QTreeWidget()
        self.resp_tree.setHeaderLabel("Responsibilities")
        self.resp_tree.itemSelectionChanged.connect(self.on_resp_select)
        splitter.addWidget(self.resp_tree)

        self.case_table = QTableWidget()
        self.case_table.setColumnCount(8)
        self.case_table.setHorizontalHeaderLabels([
            "Case No", "Date Reported", "Category", "Amount", "List", "Status", "To-Do", "Actions"
        ])

        # Enable selection change to highlight responsibility
        self.case_table.itemSelectionChanged.connect(self.on_case_select)
        # Enable double-click to view case details
        self.case_table.itemDoubleClicked.connect(lambda item: self.show_case_details(item, self.list_filter_combo.currentText()))

        # Set minimum width for headers and enable proper resizing
        header = self.case_table.horizontalHeader()
        header.setMinimumSectionSize(80)  # Minimum width for each column
        header.setSectionResizeMode(QHeaderView.Interactive)  # Allow manual resizing
        header.setStretchLastSection(True)  # Last column stretches to fill remaining space

        # Set default column widths for simplified layout
        self.case_table.setColumnWidth(0, 120)  # Case No
        self.case_table.setColumnWidth(1, 140)  # Date Reported
        self.case_table.setColumnWidth(2, 150)  # Category
        self.case_table.setColumnWidth(3, 120)  # Amount
        self.case_table.setColumnWidth(4, 120)  # List
        self.case_table.setColumnWidth(5, 120)  # Status
        self.case_table.setColumnWidth(6, 80)   # To-Do
        self.case_table.setColumnWidth(7, 80)   # Actions

        # Set row height for better readability
        self.case_table.verticalHeader().setDefaultSectionSize(25)

        splitter.addWidget(self.case_table)

        splitter.setSizes([300, 700])
        content_layout.addWidget(splitter)
        layout.addLayout(content_layout)
        self.refresh_responsibilities()
        self.refresh_cases()

    def refresh_responsibilities(self):
        self.resp_tree.clear()
        resp_dict = {r["id"]: r for r in self.responsibilities}

        # Query database to find responsibilities with cases
        self.responsibilities_with_cases = self.get_responsibilities_with_cases()

        top_level = [r for r in self.responsibilities if r["parent_id"] is None]
        for resp in top_level:
            self.add_resp_item(resp, None, resp_dict)

    def get_responsibilities_with_cases(self):
        """Get set of responsibility IDs that have cases, including their parents"""
        responsibilities_with_cases = set()
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            # Build query with financial year filter (but not for "All Cases" list filter)
            query = "SELECT DISTINCT responsibility_id FROM cases WHERE list != 'Deleted Cases'"
            params = []

            # Add financial year filter if selected
            selected_fy_id = self.fy_filter_combo.currentData()
            if selected_fy_id:
                query += " AND fy_id = ?"
                params.append(selected_fy_id)

            cursor.execute(query, params)
            case_resp_ids = {row[0] for row in cursor.fetchall()}

            # Include parent responsibilities
            for resp_id in case_resp_ids:
                responsibilities_with_cases.add(resp_id)
                # Find and add parent IDs
                resp = next((r for r in self.responsibilities if r["id"] == resp_id), None)
                if resp and resp["parent_id"]:
                    responsibilities_with_cases.add(resp["parent_id"])

            conn.close()
        except sqlite3.Error as e:
            print(f"Error querying responsibilities with cases: {e}")

        return responsibilities_with_cases

    def add_resp_item(self, resp, parent_item, resp_dict):
        item = QTreeWidgetItem([resp["name"]])
        item.setData(0, Qt.UserRole, resp["id"])

        # Bold responsibilities that have cases
        if resp["id"] in self.responsibilities_with_cases:
            font = item.font(0)
            font.setBold(True)
            item.setFont(0, font)

        if parent_item is None:
            self.resp_tree.addTopLevelItem(item)
        else:
            parent_item.addChild(item)
        children = [r for r in self.responsibilities if r["parent_id"] == resp["id"]]
        for child in children:
            self.add_resp_item(child, item, resp_dict)

    def on_resp_select(self):
        selected = self.resp_tree.selectedItems()
        if selected:
            resp_id = selected[0].data(0, Qt.UserRole)
            subtree_ids = get_subtree_resp_ids(resp_id, self.responsibilities)
            self.refresh_cases(subtree_ids)
        else:
            self.refresh_cases()

    def on_case_select(self):
        """Highlight the responsibility in the tree when a case is selected"""
        selected_rows = set()
        for item in self.case_table.selectedItems():
            selected_rows.add(item.row())

        if not selected_rows:
            # Clear selection if no case is selected
            self.resp_tree.clearSelection()
            return

        # Get the first selected case's transaction number
        first_row = min(selected_rows)
        transaction_no = self.case_table.item(first_row, 0).data(Qt.UserRole)

        # Get responsibility_id for this case
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            cursor.execute("SELECT responsibility_id FROM cases WHERE transaction_no = ?", (transaction_no,))
            result = cursor.fetchone()
            conn.close()

            if result:
                responsibility_id = result[0]
                self.highlight_responsibility(responsibility_id)
        except sqlite3.Error as e:
            print(f"Error getting responsibility for case {transaction_no}: {e}")

    def highlight_responsibility(self, responsibility_id):
        """Find and highlight the responsibility in the tree"""
        def find_item_by_id(parent_item, target_id):
            """Recursively search for an item with the given ID"""
            if parent_item is None:
                # Search top-level items
                for i in range(self.resp_tree.topLevelItemCount()):
                    item = self.resp_tree.topLevelItem(i)
                    if item.data(0, Qt.UserRole) == target_id:
                        return item
                    # Search children
                    result = find_item_by_id(item, target_id)
                    if result:
                        return result
            else:
                # Search children of parent_item
                for i in range(parent_item.childCount()):
                    item = parent_item.child(i)
                    if item.data(0, Qt.UserRole) == target_id:
                        return item
                    # Search grandchildren
                    result = find_item_by_id(item, target_id)
                    if result:
                        return result
            return None

        # Find the responsibility item
        target_item = find_item_by_id(None, responsibility_id)

        if target_item:
            # Clear current selection
            self.resp_tree.clearSelection()
            # Select the target item
            target_item.setSelected(True)
            # Ensure it's visible
            self.resp_tree.scrollToItem(target_item)
            # Expand parent items to make it visible
            parent = target_item.parent()
            while parent:
                parent.setExpanded(True)
                parent = parent.parent()

    def refresh_cases(self, resp_ids=None):
        self.case_table.setRowCount(0)
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Build base query with list filtering
        base_conditions = ["list != 'Deleted Cases'"]
        params = []

        # Add financial year filter
        selected_fy_id = self.fy_filter_combo.currentData()
        if selected_fy_id:
            base_conditions.append("fy_id = ?")
            params.append(selected_fy_id)

        # Add list filter condition using new single-case model
        selected_list = self.list_filter_combo.currentText()
        if selected_list == "Checklist":
            # Checklist shows all cases (no additional filter)
            pass
        elif selected_list == "Lead Schedule":
            # Lead Schedule shows Confirmed cases with -LS suffix, not finalized
            base_conditions.append("assessment_status = 'Confirmed' AND suffixes LIKE '%-LS%' AND suffixes NOT LIKE '%-REC%' AND suffixes NOT LIKE '%-WO%'")
        elif selected_list == "Recovered":
            # Recovered shows cases with -REC suffix
            base_conditions.append("suffixes LIKE '%-REC%'")
        elif selected_list == "Write-Off Recommended":
            # Write-Off Recommended shows cases with -WOR suffix
            base_conditions.append("suffixes LIKE '%-WOR%'")
        elif selected_list == "Written Off":
            # Written Off shows cases with -WO suffix
            base_conditions.append("suffixes LIKE '%-WO%'")
        elif selected_list == "To-Do List":
            # Show both actual To-Do List cases and GJ cases with outstanding actions
            base_conditions.append("(list = 'To-Do List' OR bas_journal_no IS NOT NULL)")
        elif selected_list == "Deleted Cases":
            # Deleted Cases shows cases with -DEL suffix
            base_conditions.append("suffixes LIKE '%-DEL%'")
        # For "All Cases", we don't add any additional list condition

        # Add responsibility filter if provided
        if resp_ids:
            placeholders = ",".join("?" for _ in resp_ids)
            base_conditions.append(f"responsibility_id IN ({placeholders})")
            params.extend(resp_ids)

        where_clause = " AND ".join(base_conditions)
        query = f"SELECT id, transaction_no, date_reported, category, amount, assessment_status, lc_status, suffixes, bas_payment_no, bas_journal_no FROM cases WHERE {where_clause}"

        cursor.execute(query, params)
        for row_data in cursor.fetchall():
            row = self.case_table.rowCount()
            self.case_table.insertRow(row)

            # Extract data from new query structure
            case_id = row_data[0]
            transaction_no = row_data[1]
            date_reported = row_data[2]
            category = row_data[3]
            amount = row_data[4]
            assessment_status = row_data[5]
            lc_status = row_data[6]
            suffixes = row_data[7]
            bas_payment_no = row_data[8] if len(row_data) > 8 else None
            bas_journal_no = row_data[9] if len(row_data) > 9 else None

            # Handle case where transaction_no is None (migration not run or test data issue)
            if transaction_no is None:
                # Fallback to case_id as transaction number for test data
                transaction_no = str(case_id)

            # Generate display transaction number
            from scripts.Utilities.workflow_utils import get_display_transaction_no
            display_transaction_no = get_display_transaction_no(transaction_no, suffixes)

            # Determine display list and status based on current view
            if selected_list == "All Cases":
                display_list = "Checklist"  # All cases are in checklist
                display_status = assessment_status
            elif selected_list == "Lead Schedule":
                display_list = "Lead Schedule"
                display_status = lc_status or "Awaiting LC determination"
            elif selected_list == "Recovered":
                display_list = "Recovered"
                display_status = "Recovered"
            elif selected_list == "Write-Off Recommended":
                display_list = "Write-Off Recommended"
                display_status = "Write Off Recommended"
            elif selected_list == "Written Off":
                display_list = "Written Off"
                display_status = "Written Off"
            else:
                display_list = selected_list
                display_status = assessment_status

            # Set columns
            # Case No
            case_item = QTableWidgetItem(display_transaction_no)
            case_item.setData(Qt.UserRole, transaction_no)  # Store transaction_no for lookup
            self.case_table.setItem(row, 0, case_item)

            # Date Reported
            self.case_table.setItem(row, 1, QTableWidgetItem(str(date_reported) if date_reported else ""))

            # Category
            self.case_table.setItem(row, 2, QTableWidgetItem(str(category) if category else ""))

            # Amount
            amount_item = format_currency_amount(amount, right_align=True)
            self.case_table.setItem(row, 3, amount_item)

            # List
            self.case_table.setItem(row, 4, QTableWidgetItem(display_list))

            # Status
            self.case_table.setItem(row, 5, QTableWidgetItem(display_status))

            # To-Do
            todo_value = "Yes" if (bas_payment_no or bas_journal_no) else "No"
            self.case_table.setItem(row, 6, QTableWidgetItem(todo_value))

            # Add Edit button in the last column
            button = QPushButton("Edit")
            button.clicked.connect(lambda: self.edit_case(transaction_no))
            self.case_table.setCellWidget(row, 7, button)
        conn.close()

    def show_case_details(self, item, selected_list=None):
        """Show detailed case information when double-clicking a case"""
        transaction_no = item.data(Qt.UserRole)
        print(f"DEBUG: Opening case: {transaction_no}")

        case_data = None
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        # Try to find case by transaction_no first, then fallback to case_id if transaction_no is None
        if transaction_no:
            cursor.execute("SELECT * FROM cases WHERE transaction_no = ?", (transaction_no,))
        else:
            # If transaction_no is None, we need to get the case_id from the table data
            # This shouldn't happen with our fallback, but just in case
            print("DEBUG: transaction_no is None, cannot open case details")
            conn.close()
            return

        case_data = cursor.fetchone()
        print(f"DEBUG: Case data found: {case_data is not None}")

        conn.close()

        if case_data:
            # Convert to dictionary for easier handling with new schema
            columns = [desc[0] for desc in cursor.description] if cursor.description else []
            case_dict = dict(zip(columns, case_data)) if columns else {}

            # Check if case is finalized
            is_finalized = case_dict.get('is_finalized', False)

            if is_finalized:
                # Show read-only details for finalized cases
                dialog = CaseDetailsDialog(case_data, self)
                dialog.exec_()
            else:
                # Open editable dialog for non-finalized cases
                from .edit_case_dialog import EditCaseDialog
                dialog = EditCaseDialog(case_dict, self, selected_list=selected_list)
                dialog.case_modified.connect(self.refresh_cases)  # Connect refresh signal
                dialog.exec_()

    def edit_case(self, transaction_no):
        """Edit case when Edit button is clicked"""
        print(f"DEBUG: edit_case called for transaction_no: {transaction_no}")

        if transaction_no:
            # Create a mock item with the transaction_no in Qt.UserRole
            from PyQt5.QtCore import Qt
            from PyQt5.QtWidgets import QTableWidgetItem
            mock_item = QTableWidgetItem()
            mock_item.setData(Qt.UserRole, transaction_no)
            self.show_case_details(mock_item, selected_list=self.list_filter_combo.currentText())
        else:
            print("DEBUG: No transaction_no provided")

    def filter_responsibilities(self, text):
        """Filter responsibilities based on search text"""
        text = text.lower()
        if not text:
            self.refresh_responsibilities()
            return

        self.resp_tree.clear()

        # Find responsibilities that match the search text
        matching_resps = []
        parent_ids_to_include = set()

        for resp in self.responsibilities:
            if text in resp["name"].lower():
                matching_resps.append(resp)
                # Recursively collect all parent IDs up to the root
                current_parent_id = resp["parent_id"]
                while current_parent_id:
                    parent_ids_to_include.add(current_parent_id)
                    # Find the parent and get its parent_id
                    parent_resp = next((r for r in self.responsibilities if r["id"] == current_parent_id), None)
                    if parent_resp:
                        current_parent_id = parent_resp["parent_id"]
                    else:
                        current_parent_id = None

        # Include all parent responsibilities
        for resp in self.responsibilities:
            if resp["id"] in parent_ids_to_include:
                matching_resps.append(resp)

        # Remove duplicates while preserving order
        seen_ids = set()
        filtered_resps = []
        for resp in matching_resps:
            if resp["id"] not in seen_ids:
                filtered_resps.append(resp)
                seen_ids.add(resp["id"])

        # Create parent map for filtered results
        parent_map = defaultdict(list)
        for resp in filtered_resps:
            parent_map[resp["parent_id"]].append(resp)

        def add_filtered_items(parent_item, parent_id):
            for resp in sorted(parent_map[parent_id], key=lambda x: x["name"]):
                item = QTreeWidgetItem([resp["name"]])
                item.setData(0, Qt.UserRole, resp["id"])

                # Bold responsibilities that have cases
                if resp["id"] in self.responsibilities_with_cases:
                    font = item.font(0)
                    font.setBold(True)
                    item.setFont(0, font)

                if parent_id is None:
                    self.resp_tree.addTopLevelItem(item)
                else:
                    parent_item.addChild(item)
                add_filtered_items(item, resp["id"])

        add_filtered_items(None, None)
        self.resp_tree.expandAll()

    def update_write_off_buttons_visibility(self):
        """Update visibility of write-off buttons based on current list filter"""
        selected_list = self.list_filter_combo.currentText()
        show_buttons = selected_list == "Write-Off Recommended"

        self.create_submission_btn.setVisible(show_buttons)
        self.approve_submission_btn.setVisible(show_buttons)

    def create_write_off_submission(self):
        """Open dialog to create a write-off submission"""
        from .write_off_submission_dialog import WriteOffSubmissionDialog
        dialog = WriteOffSubmissionDialog(self)
        dialog.exec_()
        # Refresh the case list after creating submission
        self.refresh_cases()

    def approve_write_off_submission(self):
        """Open dialog to approve a write-off submission"""
        # Get available group IDs for approval
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        try:
            cursor.execute("""
                SELECT DISTINCT write_off_group_id
                FROM cases
                WHERE write_off_group_id IS NOT NULL AND lc_status = 'Write Off Recommended'
                ORDER BY write_off_group_id
            """)

            groups = cursor.fetchall()
            conn.close()

            if not groups:
                QMessageBox.information(self, "No Submissions", "No write-off submissions available for approval.")
                return

            # For now, just approve the first group (in a real app, you'd show a selection dialog)
            group_id = groups[0][0]

            from .write_off_submission_dialog import WriteOffApprovalDialog
            dialog = WriteOffApprovalDialog(group_id, self)
            dialog.exec_()
            # Refresh the case list after approval
            self.refresh_cases()

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load submissions: {str(e)}")

    def export_to_excel(self):
        """Export the current case list to Excel format"""
        try:
            # Check if there are cases to export - show dialog if no data
            if self.case_table.rowCount() == 0:
                QMessageBox.warning(self, "No Data", "No cases to export.")
                return

            # Get current list filter for filename - replace spaces with underscores for valid filename
            current_list = self.list_filter_combo.currentText().replace(" ", "_")

            # Create year folder if it doesn't exist - ensures export directory structure
            from scripts.Utilities.financial_utils import create_year_folder
            year_folder = create_year_folder(get_financial_year())
            export_dir = os.path.join(year_folder, "Exports")
            os.makedirs(export_dir, exist_ok=True)

            # Generate filename with timestamp - format: List_Export_{list_name}_{timestamp}.xlsx
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"List_Export_{current_list}_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)

            # Collect data from table - extract headers and row data for DataFrame creation
            data = []
            headers = []

            # Get headers from table horizontal header
            for col in range(self.case_table.columnCount()):
                header_item = self.case_table.horizontalHeaderItem(col)
                if header_item:
                    headers.append(header_item.text())

            # Get data rows - iterate through all table rows and columns
            for row in range(self.case_table.rowCount()):
                row_data = {}
                for col in range(self.case_table.columnCount()):
                    item = self.case_table.item(row, col)
                    if item:
                        # Handle special case for Case No (extract transaction number from Qt.UserRole)
                        if col == 0:  # Case No column
                            transaction_no = item.data(Qt.UserRole)
                            row_data[headers[col]] = transaction_no if transaction_no else item.text()
                        else:
                            row_data[headers[col]] = item.text()
                    else:
                        # Check for widget (like buttons) - extract text if available
                        widget = self.case_table.cellWidget(row, col)
                        if widget and hasattr(widget, 'text'):
                            row_data[headers[col]] = widget.text()
                        else:
                            row_data[headers[col]] = ""

                data.append(row_data)

            # Create DataFrame and export to Excel - use pandas for data manipulation and openpyxl for Excel formatting
            import pandas as pd

            df = pd.DataFrame(data)

            # Calculate totals for Amount column if it exists - parse currency strings and sum
            total_amount = 0.0
            if 'Amount' in df.columns:
                for amount_str in df['Amount']:
                    try:
                        # Remove currency formatting (R prefix, commas) for numeric conversion
                        clean_amount = amount_str.replace('R ', '').replace(',', '').strip()
                        total_amount += float(clean_amount)
                    except (ValueError, AttributeError):
                        pass

            # Create Excel writer with openpyxl engine for advanced formatting capabilities
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write DataFrame to Excel with specified sheet name, excluding DataFrame index
                df.to_excel(writer, sheet_name=f'{current_list} Cases', index=False)

                # Get workbook and worksheet references for post-processing formatting
                workbook = writer.book
                worksheet = writer.sheets[f'{current_list} Cases']

                # Format amount column as South African currency if it exists (R with commas and 2 decimals)
                amount_col = None
                for col_num, column_title in enumerate(df.columns, 1):
                    if column_title == 'Amount':
                        amount_col = col_num
                        break

                if amount_col:
                    from openpyxl.styles import NamedStyle
                    currency_style = NamedStyle(name='currency', number_format='R #,##0.00')
                    workbook.add_named_style(currency_style)

                    # Apply currency formatting to data rows (skip header row)
                    for row_num in range(2, len(df) + 2):  # Start from row 2 (after header)
                        cell = worksheet.cell(row=row_num, column=amount_col)
                        cell.style = 'currency'

                # Add summary information at the top of the worksheet
                worksheet.insert_rows(1)
                worksheet['A1'] = f'{current_list} Cases Export'
                worksheet['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
                worksheet['A3'] = f'Total Cases: {len(df)}'

                if total_amount > 0:
                    worksheet['A4'] = f'Total Amount: R {total_amount:,.2f}'

                # Merge cells for title row spanning all columns
                from openpyxl.utils import get_column_letter
                last_col = get_column_letter(len(df.columns))
                worksheet.merge_cells(f'A1:{last_col}1')

                # Auto-adjust column widths based on content length for better readability
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter if hasattr(column[0], 'column_letter') else None
                    if column_letter is None:
                        # Handle merged cells by getting column letter from coordinate
                        for cell in column:
                            if hasattr(cell, 'column_letter'):
                                column_letter = cell.column_letter
                                break
                        if column_letter is None:
                            continue  # Skip if we can't determine column letter

                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters for readability
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            # Show success message with file details
            QMessageBox.information(
                self, "Export Successful",
                f"Case list exported successfully!\n\n"
                f"File: {filename}\n"
                f"Location: {export_dir}\n"
                f"Cases exported: {len(df)}"
            )

        except ImportError:
            # Handle missing pandas/openpyxl dependencies with user-friendly error
            QMessageBox.critical(
                self, "Missing Dependencies",
                "Excel export requires pandas and openpyxl.\n\n"
                "Please install with: pip install pandas openpyxl"
            )
        except Exception as e:
            # Catch any other export errors and show to user
            QMessageBox.critical(self, "Export Error", f"Failed to export to Excel: {str(e)}")


class CaseDetailsDialog(QDialog):
    def __init__(self, case_data, parent=None):
        super().__init__(parent)
        self.case_data = case_data
        self.setWindowTitle(f"Case Details - {case_data[1]}")  # case_data[1] is transaction_no
        self.setFixedSize(1000, 700)
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)

        # Create scroll area for case details
        scroll_area = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QFormLayout(scroll_widget)

        # Case Information Section
        case_info_group = QGroupBox("Case Information")
        case_info_layout = QFormLayout(case_info_group)

        case_info_layout.addRow("Case No:", QLabel(self.case_data[1]))
        case_info_layout.addRow("Date Incurred:", QLabel(self.case_data[2] if self.case_data[2] else "N/A"))
        case_info_layout.addRow("Date Identified:", QLabel(self.case_data[3] if self.case_data[3] else "N/A"))
        case_info_layout.addRow("Date Reported:", QLabel(self.case_data[4] if self.case_data[4] else "N/A"))
        case_info_layout.addRow("Category:", QLabel(self.case_data[9] if self.case_data[9] else "N/A"))
        case_info_layout.addRow("Amount:", QLabel(format_currency_amount(self.case_data[11]) if self.case_data[11] else "N/A"))
        case_info_layout.addRow("List:", QLabel(self.case_data[16] if self.case_data[16] else "N/A"))
        case_info_layout.addRow("Status:", QLabel(self.case_data[15] if self.case_data[15] else "N/A"))

        scroll_layout.addRow(case_info_group)

        # Description Section
        if self.case_data[5]:  # description
            desc_group = QGroupBox("Description")
            desc_layout = QVBoxLayout(desc_group)
            desc_text = QTextEdit()
            desc_text.setPlainText(self.case_data[5])
            desc_text.setReadOnly(True)
            desc_text.setMaximumHeight(100)
            desc_layout.addWidget(desc_text)
            scroll_layout.addRow(desc_group)

        # Financial Information Section
        financial_group = QGroupBox("Financial Information")
        financial_layout = QFormLayout(financial_group)

        financial_layout.addRow("BAS Payment No:", QLabel(self.case_data[6] if self.case_data[6] else "N/A"))
        financial_layout.addRow("BAS Payment Date:", QLabel(self.case_data[7] if self.case_data[7] else "N/A"))
        financial_layout.addRow("BAS Journal No:", QLabel(self.case_data[29] if len(self.case_data) > 29 and self.case_data[29] else "N/A"))
        financial_layout.addRow("BAS Journal Date:", QLabel(self.case_data[30] if len(self.case_data) > 30 and self.case_data[30] else "N/A"))
        financial_layout.addRow("Persal No:", QLabel(self.case_data[8] if self.case_data[8] else "N/A"))

        scroll_layout.addRow(financial_group)

        # Assessment Information Section
        if self.case_data[18] or self.case_data[19]:  # assessment_assessed_by or assessment_date
            assessment_group = QGroupBox("Assessment Information")
            assessment_layout = QFormLayout(assessment_group)

            assessment_layout.addRow("Assessed By:", QLabel(self.case_data[18] if self.case_data[18] else "N/A"))
            assessment_layout.addRow("Assessment Date:", QLabel(self.case_data[19] if self.case_data[19] else "N/A"))

            scroll_layout.addRow(assessment_group)

        # Additional Information Section
        additional_group = QGroupBox("Additional Information")
        additional_layout = QFormLayout(additional_group)

        additional_layout.addRow("Criminal Charges:", QLabel(self.case_data[22] if len(self.case_data) > 22 and self.case_data[22] else "N/A"))
        additional_layout.addRow("Disciplinary Process:", QLabel(self.case_data[23] if len(self.case_data) > 23 and self.case_data[23] else "N/A"))
        additional_layout.addRow("Loss Recovery:", QLabel(self.case_data[24] if len(self.case_data) > 24 and self.case_data[24] else "N/A"))

        scroll_layout.addRow(additional_group)

        # Prevention Steps Section
        if len(self.case_data) > 25 and self.case_data[25]:  # prevention_steps
            prevention_group = QGroupBox("Prevention Steps")
            prevention_layout = QVBoxLayout(prevention_group)
            prevention_text = QTextEdit()
            prevention_text.setPlainText(self.case_data[25])
            prevention_text.setReadOnly(True)
            prevention_text.setMaximumHeight(100)
            prevention_layout.addWidget(prevention_text)
            scroll_layout.addRow(prevention_group)

        scroll_area.setWidget(scroll_widget)
        scroll_area.setWidgetResizable(True)
        layout.addWidget(scroll_area)

        # Close button
        button_layout = QHBoxLayout()
        close_button = create_professional_button("Close", 'secondary')
        close_button.clicked.connect(self.accept)
        button_layout.addStretch()
        button_layout.addWidget(close_button)
        layout.addLayout(button_layout)